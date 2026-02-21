from flask import Flask, request, jsonify, send_from_directory, send_file, make_response, abort
from flask_jwt_extended import JWTManager, create_access_token, jwt_required, get_jwt_identity
from flask_cors import CORS
import bcrypt, json, os, shutil, logging, secrets, re, mimetypes
from pathlib import Path
from datetime import datetime, timedelta
from werkzeug.utils import secure_filename
from PIL import Image
import io

# Load .env file if present (development convenience)
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

try:
    import mammoth
    MAMMOTH_AVAILABLE = True
except ImportError:
    MAMMOTH_AVAILABLE = False
    print("⚠️  mammoth not installed - DOCX preview disabled. Run: pip install mammoth")

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️  openpyxl not installed - Excel preview disabled. Run: pip install openpyxl")

BASE_DIR = Path(__file__).parent
STORAGE_DIR = BASE_DIR / "nas_storage"
USERS_FILE = BASE_DIR / "users.json"
SHARED_FILE = BASE_DIR / "shared_files.json"
LOG_DIR = BASE_DIR / "logs"
STATIC_DIR = BASE_DIR / "static"

STORAGE_DIR.mkdir(exist_ok=True)
LOG_DIR.mkdir(exist_ok=True)
STATIC_DIR.mkdir(exist_ok=True)

JWT_SECRET = os.getenv("JWT_SECRET_KEY", secrets.token_urlsafe(64))

DANGEROUS_EXTENSIONS = {
    'exe', 'bat', 'cmd', 'com', 'pif', 'scr', 'vbs', 'msi', 'hta',
    'sh', 'bash', 'zsh',
    'py', 'rb', 'pl', 'php', 'cgi',
    'ps1', 'psm1', 'psd1',
    'apk', 'ipa',
}

PREVIEWABLE_TYPES = {
    'image': {'png', 'jpg', 'jpeg', 'gif', 'bmp', 'webp', 'svg', 'tiff', 'tif'},
    'pdf': {'pdf'},
    'text': {'txt', 'md', 'json', 'xml', 'csv', 'log', 'yaml', 'yml',
             'toml', 'ini', 'cfg', 'conf'},
    'docx': {'docx'},
    'xlsx': {'xlsx', 'xls'},
    'video': {'mp4', 'webm', 'ogg', 'mov', 'mkv', 'avi', 'm4v'},
}


def allowed_file(filename):
    basename = filename.replace('\\', '/').rstrip('/')
    if '/' in basename:
        basename = basename.rsplit('/', 1)[1]
    name_part = basename.lstrip('.')
    if '.' not in name_part:
        return True
    ext = name_part.rsplit('.', 1)[1].lower()
    return ext not in DANGEROUS_EXTENSIONS


def get_file_type(filename):
    if '.' not in filename:
        return 'unknown'
    ext = filename.rsplit('.', 1)[1].lower()
    for file_type, extensions in PREVIEWABLE_TYPES.items():
        if ext in extensions:
            return file_type
    return 'other'


def validate_username(username):
    if not username or len(username) < 3 or len(username) > 32:
        return False
    return bool(re.match(r'^[a-zA-Z0-9_]+$', username))


def validate_path(username, path):
    """Validate and sanitize a path, preserving hidden folder dots."""
    if not path:
        return ""
    path = path.strip().strip('/')
    # Split into components and sanitize each, preserving leading dots for hidden dirs
    components = []
    for p in path.split('/'):
        if not p or p == '..':
            continue
        # Check if it's a hidden entry (starts with dot)
        leading_dot = p.startswith('.')
        inner = p.lstrip('.')
        safe = secure_filename(inner) if inner else ''
        if not safe and not leading_dot:
            continue
        if leading_dot and safe:
            components.append('.' + safe)
        elif leading_dot and not safe:
            # e.g. ".hidden" where inner part is empty after strip — keep as-is if safe chars
            safe_raw = re.sub(r'[^\w\-]', '_', p).strip('_') or 'hidden'
            components.append(safe_raw)
        else:
            components.append(safe)
    return '/'.join(components)


def sanitize_relative_path(rel_path):
    if not rel_path:
        return ""
    rel_path = rel_path.replace("\\", "/").strip("/")
    clean_parts = []
    for part in rel_path.split("/"):
        if not part or part in (".", ".."):
            continue
        leading_dot = part.startswith('.')
        inner = part.lstrip('.')
        safe = secure_filename(inner) if inner else ''
        if leading_dot and safe:
            safe = '.' + safe
        elif leading_dot and not safe:
            safe = re.sub(r'[^\w\-.]', '_', part).strip('._') or 'file'
        elif not safe:
            safe = re.sub(r'[^\w\-.]', '_', part).strip('._') or 'file'
        clean_parts.append(safe)
    return "/".join(clean_parts)


class SharedFilesManager:
    def __init__(self, shared_file):
        self.shared_file = shared_file
        self.load_shared_files()

    def load_shared_files(self):
        if self.shared_file.exists():
            try:
                with open(self.shared_file, 'r', encoding='utf-8') as f:
                    self.shared_files = json.load(f)
            except (json.JSONDecodeError, UnicodeDecodeError):
                self.shared_files = {"pending": [], "approved": [], "folders": []}
                self.save_shared_files()
        else:
            self.shared_files = {"pending": [], "approved": [], "folders": []}
            self.save_shared_files()

    def save_shared_files(self):
        with open(self.shared_file, 'w', encoding='utf-8') as f:
            json.dump(self.shared_files, f, indent=2, ensure_ascii=False)

    def request_share(self, username, filepath, filename, file_size, file_type):
        file_entry = {
            "id": secrets.token_urlsafe(16),
            "username": username,
            "filepath": filepath,
            "filename": filename,
            "file_size": file_size,
            "file_type": file_type,
            "requested_at": datetime.now().isoformat(),
            "status": "pending"
        }
        self.shared_files["pending"].append(file_entry)
        self.save_shared_files()
        return file_entry["id"]

    def request_folder_share(self, username, folder_path, folder_name):
        folder_entry = {
            "id": secrets.token_urlsafe(16),
            "username": username,
            "folder_path": folder_path,
            "folder_name": folder_name,
            "requested_at": datetime.now().isoformat(),
            "status": "pending"
        }
        if "pending_folders" not in self.shared_files:
            self.shared_files["pending_folders"] = []
        self.shared_files["pending_folders"].append(folder_entry)
        self.save_shared_files()
        return folder_entry["id"]

    def approve_share(self, file_id):
        for i, entry in enumerate(self.shared_files["pending"]):
            if entry["id"] == file_id:
                entry["status"] = "approved"
                entry["approved_at"] = datetime.now().isoformat()
                self.shared_files["approved"].append(entry)
                self.shared_files["pending"].pop(i)
                self.save_shared_files()
                return True
        return False

    def approve_folder_share(self, folder_id):
        if "pending_folders" not in self.shared_files:
            self.shared_files["pending_folders"] = []
        if "folders" not in self.shared_files:
            self.shared_files["folders"] = []

        for i, entry in enumerate(self.shared_files["pending_folders"]):
            if entry["id"] == folder_id:
                entry["status"] = "approved"
                entry["approved_at"] = datetime.now().isoformat()
                self.shared_files["folders"].append(entry)
                self.shared_files["pending_folders"].pop(i)
                self.save_shared_files()
                return True
        return False

    def reject_share(self, file_id):
        for i, entry in enumerate(self.shared_files["pending"]):
            if entry["id"] == file_id:
                self.shared_files["pending"].pop(i)
                self.save_shared_files()
                return True
        return False

    def reject_folder_share(self, folder_id):
        if "pending_folders" not in self.shared_files:
            return False
        for i, entry in enumerate(self.shared_files["pending_folders"]):
            if entry["id"] == folder_id:
                self.shared_files["pending_folders"].pop(i)
                self.save_shared_files()
                return True
        return False

    def remove_share(self, file_id):
        for i, entry in enumerate(self.shared_files["approved"]):
            if entry["id"] == file_id:
                self.shared_files["approved"].pop(i)
                self.save_shared_files()
                return True
        return False

    def remove_folder_share(self, folder_id):
        if "folders" not in self.shared_files:
            return False
        for i, entry in enumerate(self.shared_files["folders"]):
            if entry["id"] == folder_id:
                self.shared_files["folders"].pop(i)
                self.save_shared_files()
                return True
        return False

    def get_pending(self):
        return self.shared_files["pending"]

    def get_pending_folders(self):
        if "pending_folders" not in self.shared_files:
            self.shared_files["pending_folders"] = []
        return self.shared_files["pending_folders"]

    def get_approved(self):
        return self.shared_files["approved"]

    def get_approved_folders(self):
        if "folders" not in self.shared_files:
            self.shared_files["folders"] = []
        return self.shared_files["folders"]

    def is_file_shared(self, username, filepath):
        for entry in self.shared_files["approved"]:
            if entry["username"] == username and entry["filepath"] == filepath:
                return True
        return False

    def is_folder_shared(self, username, folder_path):
        if "folders" not in self.shared_files:
            return False
        for entry in self.shared_files["folders"]:
            if entry["username"] == username and entry["folder_path"] == folder_path:
                return True
        return False

    def is_in_shared_folder(self, username, filepath):
        if "folders" not in self.shared_files:
            return False
        for folder in self.shared_files["folders"]:
            if folder["username"] == username:
                if filepath.startswith(folder["folder_path"] + "/") or filepath == folder["folder_path"]:
                    return True
        return False

    def get_shared_folder_for_path(self, username, filepath):
        """Return the shared folder entry that contains the given filepath, or None."""
        if "folders" not in self.shared_files:
            return None
        for folder in self.shared_files["folders"]:
            if folder["username"] == username:
                if filepath.startswith(folder["folder_path"] + "/") or filepath == folder["folder_path"]:
                    return folder
        return None

    def cleanup_missing_items(self):
        changed = False

        approved = self.shared_files.get("approved", [])
        cleaned_approved = []
        for entry in approved:
            user_dir = STORAGE_DIR / entry["username"]
            file_path = user_dir / entry["filepath"]
            if file_path.exists():
                cleaned_approved.append(entry)
            else:
                changed = True
                logger.info(f"Removed missing shared file: {entry['username']}/{entry['filepath']}")

        if len(cleaned_approved) != len(approved):
            self.shared_files["approved"] = cleaned_approved

        folders = self.shared_files.get("folders", [])
        cleaned_folders = []
        for entry in folders:
            user_dir = STORAGE_DIR / entry["username"]
            folder_path = user_dir / entry["folder_path"]
            if folder_path.exists() and folder_path.is_dir():
                cleaned_folders.append(entry)
            else:
                changed = True
                logger.info(f"Removed missing shared folder: {entry['username']}/{entry['folder_path']}")

        if len(cleaned_folders) != len(folders):
            self.shared_files["folders"] = cleaned_folders

        if changed:
            self.save_shared_files()

        return changed


class UserManager:
    def __init__(self, users_file):
        self.users_file = users_file
        self.load_users()

    def load_users(self):
        if self.users_file.exists():
            try:
                with open(self.users_file, 'r', encoding='utf-8') as f:
                    self.users = json.load(f)
            except (json.JSONDecodeError, UnicodeDecodeError):
                backup = self.users_file.with_suffix('.json.backup')
                self.users_file.rename(backup)
                print(f"⚠️  Corrupted users.json backed up to {backup}")
                self._create_default_user()
        else:
            self._create_default_user()

    def _create_default_user(self):
        self.users = {
            "admin": {
                "password_hash": bcrypt.hashpw("admin123".encode(), bcrypt.gensalt()).decode(),
                "role": "admin",
                "status": "approved",
                "created_at": datetime.now().isoformat(),
                "storage_used": 0,
                "trusted_uploader": True,
                "auto_share": False
            }
        }
        self.save_users()

    def save_users(self):
        with open(self.users_file, 'w', encoding='utf-8') as f:
            json.dump(self.users, f, indent=2, ensure_ascii=False)

    def register_user(self, username, password):
        if not validate_username(username):
            return False, "Invalid username format"
        if username in self.users:
            return False, "Username already exists"
        if len(password) < 6:
            return False, "Password must be at least 6 characters"

        self.users[username] = {
            "password_hash": bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode(),
            "role": "user",
            "status": "pending",
            "created_at": datetime.now().isoformat(),
            "storage_used": 0,
            "trusted_uploader": False,
            "auto_share": False
        }
        self.save_users()
        return True, "Registration submitted for approval"

    def approve_user(self, username):
        if username not in self.users:
            return False, "User not found"
        self.users[username]["status"] = "approved"
        self.save_users()
        (STORAGE_DIR / username).mkdir(exist_ok=True)
        return True, "User approved"

    def reject_user(self, username):
        if username not in self.users or self.users[username]["status"] == "approved":
            return False, "Cannot reject"
        del self.users[username]
        self.save_users()
        return True, "User rejected"

    def add_user(self, username, password, role="user"):
        if not validate_username(username):
            return False, "Invalid username format"
        if username in self.users:
            return False, "User exists"
        if len(password) < 6:
            return False, "Password must be at least 6 characters"

        self.users[username] = {
            "password_hash": bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode(),
            "role": role,
            "status": "approved",
            "created_at": datetime.now().isoformat(),
            "storage_used": 0,
            "trusted_uploader": role == "admin",
            "auto_share": False
        }
        self.save_users()
        (STORAGE_DIR / username).mkdir(exist_ok=True)
        return True, "User created"

    def delete_user(self, username):
        if username == "admin" or username not in self.users:
            return False, "Cannot delete"
        user_dir = STORAGE_DIR / username
        if user_dir.exists():
            shutil.rmtree(user_dir)
        del self.users[username]
        self.save_users()
        return True, "User deleted"

    def authenticate(self, username, password):
        if username not in self.users:
            return False
        if self.users[username]["status"] != "approved":
            return False
        return bcrypt.checkpw(password.encode(), self.users[username]["password_hash"].encode())

    def get_user(self, username):
        return self.users.get(username)

    def list_users(self, status=None):
        return [
            {
                "username": u,
                "role": d["role"],
                "status": d["status"],
                "created_at": d["created_at"],
                "storage_used": d["storage_used"],
                "trusted_uploader": d.get("trusted_uploader", False),
                "auto_share": d.get("auto_share", False)
            }
            for u, d in self.users.items()
            if status is None or d["status"] == status
        ]

    def update_storage_used(self, username):
        user_dir = STORAGE_DIR / username
        if user_dir.exists():
            total = sum(f.stat().st_size for f in user_dir.rglob('*') if f.is_file())
            self.users[username]["storage_used"] = total
            self.save_users()

    def set_trusted_uploader(self, username, value: bool):
        if username not in self.users:
            return False, "User not found"
        self.users[username]["trusted_uploader"] = value
        self.save_users()
        return True, "Updated"

    def set_auto_share(self, username, value: bool):
        if username not in self.users:
            return False, "User not found"
        self.users[username]["auto_share"] = value
        self.save_users()
        return True, "Updated"


app = Flask(__name__, static_folder='static')
app.config.update(
    JWT_SECRET_KEY=JWT_SECRET,
    JWT_ACCESS_TOKEN_EXPIRES=timedelta(hours=8),
    JWT_TOKEN_LOCATION=["headers", "query_string"],
    JWT_QUERY_STRING_NAME="token",
    MAX_CONTENT_LENGTH=2 * 1024 * 1024 * 1024,  # 2 GB
)
CORS(app)
jwt = JWTManager(app)
user_manager = UserManager(USERS_FILE)
shared_manager = SharedFilesManager(SHARED_FILE)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOG_DIR / 'nas.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


# ==================== GLOBAL ERROR HANDLERS ====================

@app.errorhandler(400)
def bad_request(e):
    return jsonify({"msg": "Bad request", "error": str(e)}), 400

@app.errorhandler(401)
def unauthorized(e):
    return jsonify({"msg": "Unauthorized"}), 401

@app.errorhandler(403)
def forbidden(e):
    return jsonify({"msg": "Forbidden"}), 403

@app.errorhandler(404)
def not_found(e):
    if request.path.startswith('/api/'):
        return jsonify({"msg": f"Endpoint not found: {request.path}"}), 404
    return send_file(STATIC_DIR / "index.html")

@app.errorhandler(405)
def method_not_allowed(e):
    return jsonify({"msg": "Method not allowed"}), 405

@app.errorhandler(500)
def internal_error(e):
    logger.error(f"Internal server error: {e}")
    return jsonify({"msg": "Internal server error", "error": str(e)}), 500

@app.errorhandler(Exception)
def unhandled_exception(e):
    logger.error(f"Unhandled exception: {e}")
    return jsonify({"msg": "Unexpected server error", "error": str(e)}), 500


def format_size(size):
    for unit in ['B', 'KB', 'MB', 'GB', 'TB']:
        if size < 1024:
            return f"{size:.2f} {unit}"
        size /= 1024
    return f"{size:.2f} PB"


def should_auto_approve(user):
    """Returns True if a share request from this user should be auto-approved."""
    return user.get("role") == "admin" or user.get("trusted_uploader", False)


# ==================== STATIC FILES ====================
@app.route("/")
def index():
    return send_file(STATIC_DIR / "index.html")


@app.route("/static/<path:filename>")
def serve_static(filename):
    return send_from_directory(STATIC_DIR, filename)


# ==================== API ENDPOINTS ====================
@app.route("/api/register", methods=["POST"])
def register():
    data = request.get_json()
    username = data.get("username", "").strip()
    password = data.get("password", "")

    success, msg = user_manager.register_user(username, password)

    if success:
        logger.info(f"New user registration: {username}")
        return jsonify({"msg": msg}), 200
    return jsonify({"msg": msg}), 400


@app.route("/api/login", methods=["POST"])
def login():
    data = request.get_json()
    username = data.get("username", "").strip()
    password = data.get("password", "")

    if not user_manager.authenticate(username, password):
        logger.warning(f"Failed login attempt for: {username}")
        return jsonify({"msg": "Invalid credentials"}), 401

    user = user_manager.get_user(username)
    access_token = create_access_token(identity=username)

    logger.info(f"User logged in: {username}")
    return jsonify({
        "access_token": access_token,
        "role": user["role"],
        "trusted_uploader": user.get("trusted_uploader", False),
        "msg": "Login successful"
    }), 200


@app.route("/api/upload", methods=["POST"])
@jwt_required()
def upload_files():
    username = get_jwt_identity()
    user = user_manager.get_user(username)
    folder_path = request.form.get('folder', '').strip()
    folder_path = validate_path(username, folder_path)
    is_folder_upload = request.form.get('is_folder_upload', 'false') == 'true'

    user_dir = STORAGE_DIR / username
    base_dir = user_dir / folder_path if folder_path else user_dir
    base_dir.mkdir(parents=True, exist_ok=True)

    if 'files' not in request.files:
        return jsonify({"msg": "No files provided"}), 400

    files = request.files.getlist('files')
    paths = request.form.getlist('paths')

    uploaded = []
    errors = []

    # Check if uploading into an already-shared folder
    shared_folder_entry = shared_manager.get_shared_folder_for_path(username, folder_path) if folder_path else None

    for idx, file in enumerate(files):
        if file.filename == '':
            continue

        if not allowed_file(file.filename):
            errors.append(f"{file.filename}: File type not allowed")
            continue

        if is_folder_upload and idx < len(paths):
            rel_path = sanitize_relative_path(paths[idx])
            if not rel_path:
                rel_path = secure_filename(file.filename)
            target_path = base_dir / rel_path
            target_path.parent.mkdir(parents=True, exist_ok=True)
            if target_path.exists():
                name, ext = os.path.splitext(target_path.name)
                counter = 1
                while target_path.exists():
                    target_path = target_path.parent / f"{name}_{counter}{ext}"
                    counter += 1
        else:
            filename = secure_filename(file.filename)
            target_path = base_dir / filename
            if target_path.exists():
                name, ext = os.path.splitext(filename)
                counter = 1
                while target_path.exists():
                    target_path = base_dir / f"{name}_{counter}{ext}"
                    counter += 1

        try:
            file.save(str(target_path))
            rel = str(target_path.relative_to(user_dir))
            uploaded.append(rel)

            # Auto-share if: user has auto_share flag, OR uploading into a shared folder
            if user.get("auto_share", False) or shared_folder_entry:
                stat = target_path.stat()
                file_id = shared_manager.request_share(
                    username, rel, target_path.name,
                    stat.st_size, get_file_type(target_path.name)
                )
                shared_manager.approve_share(file_id)

            logger.info(f"File uploaded: {username}/{rel}")
        except Exception as e:
            logger.error(f"Upload error for {username}/{file.filename}: {str(e)}")
            errors.append(f"{file.filename}: Upload failed")

    user_manager.update_storage_used(username)

    if uploaded:
        msg = f"Uploaded {len(uploaded)} file(s)"
        if errors:
            msg += f". {len(errors)} file(s) failed"
        if user.get("auto_share", False) or shared_folder_entry:
            msg += " and auto-shared to network"
        return jsonify({"msg": msg, "uploaded": uploaded}), 200

    return jsonify({"msg": "No files uploaded", "errors": errors}), 400


@app.route("/api/stats", methods=["GET"])
@jwt_required()
def get_stats():
    username = get_jwt_identity()
    user_dir = STORAGE_DIR / username

    if not user_dir.exists():
        return jsonify({"total_files": 0, "total_size": 0, "total_size_formatted": "0 B"}), 200

    total_files = sum(1 for f in user_dir.rglob('*') if f.is_file())
    total_size  = sum(f.stat().st_size for f in user_dir.rglob('*') if f.is_file())

    return jsonify({
        "total_files": total_files,
        "total_size": total_size,
        "total_size_formatted": format_size(total_size)
    }), 200


@app.route("/api/files", methods=["GET"])
@jwt_required()
def list_files():
    username = get_jwt_identity()
    folder_path = request.args.get('folder', '').strip()
    folder_path = validate_path(username, folder_path)

    user_dir = STORAGE_DIR / username
    if folder_path:
        target_dir = user_dir / folder_path
    else:
        target_dir = user_dir

    if not target_dir.exists():
        return jsonify({
            "files": [],
            "folders": [],
            "current_path": folder_path,
            "total_files": 0,
            "total_size": 0,
            "total_size_formatted": "0 B"
        }), 200

    files = []
    folders = []
    total_size = 0

    for item in target_dir.iterdir():
        if item.is_file():
            stat = item.stat()
            ext = item.suffix[1:].lower() if item.suffix else ''
            filepath_relative = str(item.relative_to(user_dir))
            is_shared = shared_manager.is_file_shared(username,
                                                      filepath_relative) or shared_manager.is_in_shared_folder(username,
                                                                                                               filepath_relative)
            files.append({
                "name": item.name,
                "size": stat.st_size,
                "size_formatted": format_size(stat.st_size),
                "modified": stat.st_mtime,
                "type": get_file_type(item.name),
                "ext": ext,
                "is_shared": is_shared
            })
            total_size += stat.st_size
        elif item.is_dir():
            stat = item.stat()
            folder_relative = str(item.relative_to(user_dir))
            folders.append({
                "name": item.name,
                "modified": stat.st_mtime,
                "is_shared": shared_manager.is_folder_shared(username, folder_relative)
            })

    files.sort(key=lambda x: x["modified"], reverse=True)
    folders.sort(key=lambda x: x["name"])

    return jsonify({
        "files": files,
        "folders": folders,
        "current_path": folder_path,
        "total_files": len(files),
        "total_size": total_size,
        "total_size_formatted": format_size(total_size)
    }), 200


@app.route("/api/share/request", methods=["POST"])
@jwt_required()
def request_share():
    username = get_jwt_identity()
    user = user_manager.get_user(username)
    data = request.get_json()
    filepath = data.get('filepath', '').strip()

    filepath = validate_path(username, filepath)
    user_dir = STORAGE_DIR / username
    file_path = user_dir / filepath

    try:
        file_path.resolve().relative_to(user_dir.resolve())
    except ValueError:
        return jsonify({"msg": "Access denied"}), 403

    if not file_path.exists():
        return jsonify({"msg": "File not found"}), 404

    stat = file_path.stat()
    file_id = shared_manager.request_share(
        username,
        filepath,
        file_path.name,
        stat.st_size,
        get_file_type(file_path.name)
    )

    # Auto-approve for admins and trusted uploaders
    if should_auto_approve(user):
        shared_manager.approve_share(file_id)
        logger.info(f"Share auto-approved for {username}: {filepath}")
        return jsonify({"msg": "File shared to network", "file_id": file_id, "auto_approved": True}), 200

    logger.info(f"Share request: {username}/{filepath}")
    return jsonify({"msg": "Share request submitted for approval", "file_id": file_id, "auto_approved": False}), 200


@app.route("/api/share/folder/request", methods=["POST"])
@jwt_required()
def request_folder_share():
    username = get_jwt_identity()
    user = user_manager.get_user(username)
    data = request.get_json()
    folder_path = data.get('folder_path', '').strip()

    folder_path = validate_path(username, folder_path)
    user_dir = STORAGE_DIR / username
    folder_full_path = user_dir / folder_path

    try:
        folder_full_path.resolve().relative_to(user_dir.resolve())
    except ValueError:
        return jsonify({"msg": "Access denied"}), 403

    if not folder_full_path.exists() or not folder_full_path.is_dir():
        return jsonify({"msg": "Folder not found"}), 404

    folder_id = shared_manager.request_folder_share(
        username,
        folder_path,
        folder_full_path.name
    )

    # Auto-approve for admins and trusted uploaders
    if should_auto_approve(user):
        shared_manager.approve_folder_share(folder_id)
        logger.info(f"Folder share auto-approved for {username}: {folder_path}")
        return jsonify({"msg": "Folder shared to network", "folder_id": folder_id, "auto_approved": True}), 200

    logger.info(f"Folder share request: {username}/{folder_path}")
    return jsonify({"msg": "Folder share request submitted for approval", "folder_id": folder_id, "auto_approved": False}), 200


@app.route("/api/share/pending", methods=["GET"])
@jwt_required()
def get_pending_shares():
    username = get_jwt_identity()
    user = user_manager.get_user(username)

    if user["role"] != "admin":
        return jsonify({"msg": "Admin access required"}), 403

    pending_files = shared_manager.get_pending()
    pending_folders = shared_manager.get_pending_folders()
    return jsonify({"shares": pending_files, "folders": pending_folders}), 200


@app.route("/api/share/approve/<file_id>", methods=["POST"])
@jwt_required()
def approve_share(file_id):
    username = get_jwt_identity()
    user = user_manager.get_user(username)

    if user["role"] != "admin":
        return jsonify({"msg": "Admin access required"}), 403

    if shared_manager.approve_share(file_id):
        logger.info(f"Share approved by {username}: {file_id}")
        return jsonify({"msg": "Share approved"}), 200

    return jsonify({"msg": "Share not found"}), 404


@app.route("/api/share/folder/approve/<folder_id>", methods=["POST"])
@jwt_required()
def approve_folder_share(folder_id):
    username = get_jwt_identity()
    user = user_manager.get_user(username)

    if user["role"] != "admin":
        return jsonify({"msg": "Admin access required"}), 403

    if shared_manager.approve_folder_share(folder_id):
        logger.info(f"Folder share approved by {username}: {folder_id}")
        return jsonify({"msg": "Folder share approved"}), 200

    return jsonify({"msg": "Folder share not found"}), 404


@app.route("/api/share/reject/<file_id>", methods=["POST"])
@jwt_required()
def reject_share(file_id):
    username = get_jwt_identity()
    user = user_manager.get_user(username)

    if user["role"] != "admin":
        return jsonify({"msg": "Admin access required"}), 403

    if shared_manager.reject_share(file_id):
        logger.info(f"Share rejected by {username}: {file_id}")
        return jsonify({"msg": "Share rejected"}), 200

    return jsonify({"msg": "Share not found"}), 404


@app.route("/api/share/folder/reject/<folder_id>", methods=["POST"])
@jwt_required()
def reject_folder_share(folder_id):
    username = get_jwt_identity()
    user = user_manager.get_user(username)

    if user["role"] != "admin":
        return jsonify({"msg": "Admin access required"}), 403

    if shared_manager.reject_folder_share(folder_id):
        logger.info(f"Folder share rejected by {username}: {folder_id}")
        return jsonify({"msg": "Folder share rejected"}), 200

    return jsonify({"msg": "Folder share not found"}), 404


@app.route("/api/share/remove/<file_id>", methods=["DELETE"])
@jwt_required()
def remove_share(file_id):
    username = get_jwt_identity()
    user = user_manager.get_user(username)

    approved = shared_manager.get_approved()
    file_entry = None
    for entry in approved:
        if entry["id"] == file_id:
            file_entry = entry
            break

    if not file_entry:
        return jsonify({"msg": "Share not found"}), 404

    if user["role"] != "admin" and file_entry["username"] != username:
        return jsonify({"msg": "Permission denied"}), 403

    if shared_manager.remove_share(file_id):
        logger.info(f"Share removed by {username}: {file_id}")
        return jsonify({"msg": "Share removed"}), 200

    return jsonify({"msg": "Share not found"}), 404


@app.route("/api/share/folder/remove/<folder_id>", methods=["DELETE"])
@jwt_required()
def remove_folder_share(folder_id):
    username = get_jwt_identity()
    user = user_manager.get_user(username)

    approved_folders = shared_manager.get_approved_folders()
    folder_entry = None
    for entry in approved_folders:
        if entry["id"] == folder_id:
            folder_entry = entry
            break

    if not folder_entry:
        return jsonify({"msg": "Folder share not found"}), 404

    if user["role"] != "admin" and folder_entry["username"] != username:
        return jsonify({"msg": "Permission denied"}), 403

    if shared_manager.remove_folder_share(folder_id):
        logger.info(f"Folder share removed by {username}: {folder_id}")
        return jsonify({"msg": "Folder share removed"}), 200

    return jsonify({"msg": "Folder share not found"}), 404


@app.route("/api/network/files", methods=["GET"])
@jwt_required()
def get_network_files():
    """Network files - requires login. Also cleans up missing items."""
    shared_manager.cleanup_missing_items()

    approved_files = shared_manager.get_approved()
    approved_folders = shared_manager.get_approved_folders()
    return jsonify({"files": approved_files, "folders": approved_folders}), 200


@app.route("/api/network/folder/<folder_id>/delete", methods=["DELETE"])
@jwt_required()
def delete_item_in_shared_folder(folder_id):
    requester = get_jwt_identity()
    requester_user = user_manager.get_user(requester)

    approved_folders = shared_manager.get_approved_folders()
    folder_entry = next((f for f in approved_folders if f["id"] == folder_id), None)
    if not folder_entry:
        return jsonify({"msg": "Shared folder not found"}), 404

    owner = folder_entry["username"]
    if requester != owner and requester_user.get("role") != "admin":
        return jsonify({"msg": "Access denied — only the owner or an admin can delete items"}), 403

    item_rel = request.args.get("path", "").strip()
    item_type = request.args.get("type", "file").strip()

    if not item_rel:
        return jsonify({"msg": "No path provided"}), 400

    item_rel = validate_path(owner, item_rel)
    user_dir = STORAGE_DIR / owner
    base_folder = user_dir / folder_entry["folder_path"]
    item_path = base_folder / item_rel

    try:
        item_path.resolve().relative_to(base_folder.resolve())
    except ValueError:
        return jsonify({"msg": "Access denied — path escapes shared folder"}), 403

    if not item_path.exists():
        return jsonify({"msg": "Item not found"}), 404

    try:
        if item_type == "folder" and item_path.is_dir():
            shutil.rmtree(item_path)
            logger.info(f"Shared folder item deleted (dir) by {requester}: {owner}/{folder_entry['folder_path']}/{item_rel}")
        elif item_path.is_file():
            item_path.unlink()
            logger.info(f"Shared folder item deleted (file) by {requester}: {owner}/{folder_entry['folder_path']}/{item_rel}")
        else:
            return jsonify({"msg": "Type mismatch or item not found"}), 400

        user_manager.update_storage_used(owner)
        return jsonify({"msg": "Deleted successfully"}), 200
    except Exception as e:
        logger.error(f"Error deleting shared folder item: {e}")
        return jsonify({"msg": f"Delete failed: {str(e)}"}), 500


@app.route("/api/network/folder/<folder_id>", methods=["GET"])
@jwt_required()
def get_network_folder(folder_id):
    """Get contents of a shared folder with proper navigation."""
    approved_folders = shared_manager.get_approved_folders()

    folder_entry = None
    for entry in approved_folders:
        if entry["id"] == folder_id:
            folder_entry = entry
            break

    if not folder_entry:
        return jsonify({"msg": "Folder not found"}), 404

    subfolder = request.args.get('subfolder', '').strip()
    subfolder = validate_path(folder_entry["username"], subfolder)

    user_dir = STORAGE_DIR / folder_entry["username"]
    base_folder_path = user_dir / folder_entry["folder_path"]

    if not base_folder_path.exists():
        shared_manager.remove_folder_share(folder_id)
        return jsonify({"msg": "Folder no longer exists and has been removed from shares"}), 404

    if subfolder:
        current_path = base_folder_path / subfolder
    else:
        current_path = base_folder_path

    if not current_path.exists() or not current_path.is_dir():
        return jsonify({"msg": "Folder not found"}), 404

    files = []
    folders = []

    for item in current_path.iterdir():
        if item.is_file():
            stat = item.stat()
            relative_to_base = str(item.relative_to(base_folder_path))
            files.append({
                "id": secrets.token_urlsafe(16),
                "username": folder_entry["username"],
                "filepath": str(item.relative_to(user_dir)),
                "filename": item.name,
                "relative_path": relative_to_base,
                "file_size": stat.st_size,
                "file_type": get_file_type(item.name),
                "modified": stat.st_mtime
            })
        elif item.is_dir():
            stat = item.stat()
            relative_to_base = str(item.relative_to(base_folder_path))
            folders.append({
                "name": item.name,
                "relative_path": relative_to_base,
                "modified": stat.st_mtime
            })

    folders.sort(key=lambda x: x["name"])
    files.sort(key=lambda x: x["modified"], reverse=True)

    return jsonify({
        "folder": folder_entry,
        "files": files,
        "folders": folders,
        "current_subfolder": subfolder
    }), 200


@app.route("/api/folder/create", methods=["POST"])
@jwt_required()
def create_folder():
    username = get_jwt_identity()
    data = request.get_json()
    current_path = data.get('current_path', '').strip()
    folder_name = data.get('folder_name', '').strip()

    if not folder_name:
        return jsonify({"msg": "Folder name required"}), 400

    current_path = validate_path(username, current_path)

    # Preserve hidden folder dots
    leading_dot = folder_name.startswith('.')
    inner = folder_name.lstrip('.')
    safe_name = secure_filename(inner) if inner else ''
    if leading_dot and safe_name:
        folder_name = '.' + safe_name
    elif not leading_dot:
        folder_name = safe_name or folder_name

    if not folder_name:
        return jsonify({"msg": "Invalid folder name"}), 400

    user_dir = STORAGE_DIR / username
    if current_path:
        target_dir = user_dir / current_path / folder_name
    else:
        target_dir = user_dir / folder_name

    if target_dir.exists():
        return jsonify({"msg": "Folder already exists"}), 400

    try:
        target_dir.mkdir(parents=True, exist_ok=False)
        logger.info(f"Folder created: {username}/{current_path}/{folder_name}")
        return jsonify({"msg": "Folder created successfully"}), 200
    except Exception as e:
        logger.error(f"Folder creation error: {str(e)}")
        return jsonify({"msg": "Failed to create folder"}), 500


@app.route("/api/folder/delete", methods=["DELETE"])
@jwt_required()
def delete_folder():
    username = get_jwt_identity()
    folder_path = request.args.get('path', '').strip()
    folder_path = validate_path(username, folder_path)

    if not folder_path:
        return jsonify({"msg": "Invalid folder path"}), 400

    user_dir = STORAGE_DIR / username
    target_dir = user_dir / folder_path

    try:
        target_dir.resolve().relative_to(user_dir.resolve())
    except ValueError:
        return jsonify({"msg": "Access denied"}), 403

    if not target_dir.exists() or not target_dir.is_dir():
        return jsonify({"msg": "Folder not found"}), 404

    try:
        shutil.rmtree(target_dir)
        user_manager.update_storage_used(username)
        # Remove folder and all files within it from shared cache
        approved_folders = list(shared_manager.get_approved_folders())
        for entry in approved_folders:
            if entry["username"] == username and (
                entry["folder_path"] == folder_path or
                entry["folder_path"].startswith(folder_path + "/")
            ):
                shared_manager.remove_folder_share(entry["id"])
        approved = list(shared_manager.get_approved())
        for entry in approved:
            if entry["username"] == username and (
                entry["filepath"] == folder_path or
                entry["filepath"].startswith(folder_path + "/")
            ):
                shared_manager.remove_share(entry["id"])
        pending = list(shared_manager.get_pending())
        for entry in pending:
            if entry["username"] == username and entry["filepath"].startswith(folder_path + "/"):
                shared_manager.reject_share(entry["id"])
        logger.info(f"Folder deleted: {username}/{folder_path}")
        return jsonify({"msg": "Folder deleted successfully"}), 200
    except Exception as e:
        logger.error(f"Folder deletion error: {str(e)}")
        return jsonify({"msg": "Failed to delete folder"}), 500


@app.route("/api/preview/<path:filepath>", methods=["GET"])
def preview_file(filepath):
    token = request.args.get('token')
    if not token:
        return jsonify({"msg": "No token provided"}), 401

    try:
        from flask_jwt_extended import decode_token
        decoded = decode_token(token)
        username = decoded['sub']
    except Exception:
        return jsonify({"msg": "Invalid token"}), 401

    filepath = validate_path(username, filepath)
    user_dir = STORAGE_DIR / username
    file_path = user_dir / filepath

    try:
        file_path.resolve().relative_to(user_dir.resolve())
    except ValueError:
        return jsonify({"msg": "Access denied"}), 403

    if not file_path.exists():
        return jsonify({"msg": "File not found"}), 404

    file_type = get_file_type(file_path.name)

    if file_type == 'image':
        try:
            img = Image.open(file_path)
            img.thumbnail((1920, 1080), Image.Resampling.LANCZOS)
            img_io = io.BytesIO()
            img.save(img_io, format=img.format or 'PNG')
            img_io.seek(0)
            return send_file(img_io, mimetype=f'image/{img.format.lower()}')
        except:
            return send_file(file_path, mimetype=mimetypes.guess_type(file_path)[0])

    elif file_type == 'pdf':
        return send_file(file_path, mimetype='application/pdf')

    elif file_type == 'text':
        return send_file(file_path, mimetype='text/plain')

    elif file_type == 'video':
        mime = mimetypes.guess_type(file_path)[0] or 'video/mp4'
        return _stream_video(file_path, mime)

    else:
        return jsonify({"msg": "Preview not available for this file type"}), 400


def _stream_video(file_path, mime_type):
    """Stream a video file with range request support for seeking."""
    file_size = file_path.stat().st_size
    range_header = request.headers.get('Range', None)

    if range_header:
        byte_start, byte_end = 0, None
        match = re.match(r'bytes=(\d+)-(\d*)', range_header)
        if match:
            byte_start = int(match.group(1))
            byte_end = int(match.group(2)) if match.group(2) else file_size - 1
        byte_end = min(byte_end, file_size - 1)
        length = byte_end - byte_start + 1

        def generate():
            with open(file_path, 'rb') as f:
                f.seek(byte_start)
                remaining = length
                while remaining > 0:
                    chunk = f.read(min(65536, remaining))
                    if not chunk:
                        break
                    remaining -= len(chunk)
                    yield chunk

        response = app.response_class(generate(), 206, mimetype=mime_type, direct_passthrough=True)
        response.headers['Content-Range'] = f'bytes {byte_start}-{byte_end}/{file_size}'
        response.headers['Accept-Ranges'] = 'bytes'
        response.headers['Content-Length'] = length
        return response
    else:
        response = send_file(file_path, mimetype=mime_type)
        response.headers['Accept-Ranges'] = 'bytes'
        response.headers['Content-Length'] = file_size
        return response


@app.route("/api/network/preview/<file_id>", methods=["GET"])
def preview_network_file(file_id):
    """Preview for shared files - requires login via token"""
    token = request.args.get('token')
    if not token:
        return jsonify({"msg": "No token provided"}), 401

    try:
        from flask_jwt_extended import decode_token
        decoded = decode_token(token)
        username = decoded['sub']
    except Exception:
        return jsonify({"msg": "Invalid token"}), 401

    approved = shared_manager.get_approved()

    file_entry = None
    for entry in approved:
        if entry["id"] == file_id:
            file_entry = entry
            break

    if not file_entry:
        return jsonify({"msg": "File not found"}), 404

    user_dir = STORAGE_DIR / file_entry["username"]
    file_path = user_dir / file_entry["filepath"]

    if not file_path.exists():
        return jsonify({"msg": "File not found"}), 404

    file_type = get_file_type(file_path.name)

    if file_type == 'image':
        try:
            img = Image.open(file_path)
            img.thumbnail((1920, 1080), Image.Resampling.LANCZOS)
            img_io = io.BytesIO()
            img.save(img_io, format=img.format or 'PNG')
            img_io.seek(0)
            return send_file(img_io, mimetype=f'image/{img.format.lower()}')
        except:
            return send_file(file_path, mimetype=mimetypes.guess_type(file_path)[0])

    elif file_type == 'pdf':
        return send_file(file_path, mimetype='application/pdf')

    elif file_type == 'text':
        return send_file(file_path, mimetype='text/plain')

    elif file_type == 'video':
        mime = mimetypes.guess_type(file_path)[0] or 'video/mp4'
        return _stream_video(file_path, mime)

    else:
        return jsonify({"msg": "Preview not available for this file type"}), 400


@app.route("/api/download/<path:filepath>", methods=["GET"])
def download_file(filepath):
    token = request.args.get('token')
    if not token:
        return jsonify({"msg": "No token provided"}), 401

    try:
        from flask_jwt_extended import decode_token
        decoded = decode_token(token)
        username = decoded['sub']
    except Exception:
        return jsonify({"msg": "Invalid token"}), 401

    filepath = validate_path(username, filepath)
    user_dir = STORAGE_DIR / username
    file_path = user_dir / filepath

    try:
        file_path.resolve().relative_to(user_dir.resolve())
    except ValueError:
        logger.warning(f"Path traversal attempt by {username}: {filepath}")
        return jsonify({"msg": "Access denied"}), 403

    if not file_path.exists():
        return jsonify({"msg": "File not found"}), 404

    logger.info(f"File downloaded: {username}/{filepath}")
    return send_file(file_path, as_attachment=True, download_name=file_path.name)


@app.route("/api/network/download/<file_id>", methods=["GET"])
def download_network_file(file_id):
    """Download for shared files - requires login via token"""
    token = request.args.get('token')
    if not token:
        return jsonify({"msg": "No token provided"}), 401

    try:
        from flask_jwt_extended import decode_token
        decoded = decode_token(token)
        username = decoded['sub']
    except Exception:
        return jsonify({"msg": "Invalid token"}), 401

    approved = shared_manager.get_approved()

    file_entry = None
    for entry in approved:
        if entry["id"] == file_id:
            file_entry = entry
            break

    if not file_entry:
        return jsonify({"msg": "File not found"}), 404

    user_dir = STORAGE_DIR / file_entry["username"]
    file_path = user_dir / file_entry["filepath"]

    if not file_path.exists():
        return jsonify({"msg": "File not found"}), 404

    logger.info(f"Network file downloaded: {file_id} by {username}")
    return send_file(file_path, as_attachment=True, download_name=file_path.name)


@app.route("/api/delete/<path:filepath>", methods=["DELETE"])
@jwt_required()
def delete_file(filepath):
    username = get_jwt_identity()
    filepath = validate_path(username, filepath)
    user_dir = STORAGE_DIR / username
    file_path = user_dir / filepath

    try:
        file_path.resolve().relative_to(user_dir.resolve())
    except ValueError:
        logger.warning(f"Path traversal attempt by {username}: {filepath}")
        return jsonify({"msg": "Access denied"}), 403

    if not file_path.exists():
        return jsonify({"msg": "File not found"}), 404

    try:
        file_path.unlink()
        user_manager.update_storage_used(username)
        # Remove from shared cache
        approved = list(shared_manager.get_approved())
        for entry in approved:
            if entry["username"] == username and entry["filepath"] == filepath:
                shared_manager.remove_share(entry["id"])
        pending = list(shared_manager.get_pending())
        for entry in pending:
            if entry["username"] == username and entry["filepath"] == filepath:
                shared_manager.reject_share(entry["id"])
        logger.info(f"File deleted: {username}/{filepath}")
        return jsonify({"msg": "File deleted"}), 200
    except Exception as e:
        logger.error(f"Delete error for {username}/{filepath}: {str(e)}")
        return jsonify({"msg": "Delete failed"}), 500


@app.route("/api/bulk/delete", methods=["POST"])
@jwt_required()
def bulk_delete_files():
    username = get_jwt_identity()
    data = request.get_json()
    filepaths = data.get('filepaths', [])

    if not filepaths:
        return jsonify({"msg": "No files specified"}), 400

    user_dir = STORAGE_DIR / username
    deleted = []
    errors = []

    for filepath in filepaths:
        filepath = validate_path(username, filepath)
        file_path = user_dir / filepath

        try:
            file_path.resolve().relative_to(user_dir.resolve())
            if file_path.exists() and file_path.is_file():
                file_path.unlink()
                deleted.append(filepath)
                # Remove from shared cache
                for entry in list(shared_manager.get_approved()):
                    if entry["username"] == username and entry["filepath"] == filepath:
                        shared_manager.remove_share(entry["id"])
                for entry in list(shared_manager.get_pending()):
                    if entry["username"] == username and entry["filepath"] == filepath:
                        shared_manager.reject_share(entry["id"])
                logger.info(f"Bulk delete: {username}/{filepath}")
            else:
                errors.append(filepath)
        except Exception as e:
            logger.error(f"Bulk delete error for {username}/{filepath}: {str(e)}")
            errors.append(filepath)

    user_manager.update_storage_used(username)

    msg = f"Deleted {len(deleted)} file(s)"
    if errors:
        msg += f". {len(errors)} failed"

    return jsonify({"msg": msg, "deleted": deleted, "errors": errors}), 200


@app.route("/api/move", methods=["POST"])
@jwt_required()
def move_item():
    """Move a single file or folder to a new destination folder."""
    username = get_jwt_identity()
    data = request.get_json()
    src = (data.get('src') or '').strip()
    dst_folder = (data.get('dst_folder') or '').strip()

    if not src:
        return jsonify({"msg": "Source path required"}), 400

    src = validate_path(username, src)
    dst_folder = validate_path(username, dst_folder) if dst_folder else ''

    user_dir = STORAGE_DIR / username
    src_path = user_dir / src
    dst_dir  = user_dir / dst_folder if dst_folder else user_dir

    # Safety checks
    try:
        src_path.resolve().relative_to(user_dir.resolve())
        dst_dir.resolve().relative_to(user_dir.resolve())
    except ValueError:
        return jsonify({"msg": "Access denied"}), 403

    if not src_path.exists():
        return jsonify({"msg": "Source not found"}), 404
    if not dst_dir.exists():
        return jsonify({"msg": "Destination folder not found"}), 404

    # Prevent moving folder into itself
    if src_path.is_dir():
        try:
            dst_dir.resolve().relative_to(src_path.resolve())
            return jsonify({"msg": "Cannot move a folder into itself"}), 400
        except ValueError:
            pass

    dst_path = dst_dir / src_path.name
    if dst_path.resolve() == src_path.resolve():
        return jsonify({"msg": "Already in that location"}), 400

    # Avoid collision
    if dst_path.exists():
        base = src_path.stem if src_path.is_file() else src_path.name
        ext  = src_path.suffix if src_path.is_file() else ''
        counter = 1
        while dst_path.exists():
            dst_path = dst_dir / f"{base}_{counter}{ext}"
            counter += 1

    try:
        shutil.move(str(src_path), str(dst_path))
        user_manager.update_storage_used(username)
        logger.info(f"Move: {username}/{src} -> {dst_folder or '(root)'}")
        return jsonify({"msg": "Moved", "new_path": str(dst_path.relative_to(user_dir))}), 200
    except Exception as e:
        logger.error(f"Move error: {e}")
        return jsonify({"msg": f"Move failed: {str(e)}"}), 500


@app.route("/api/bulk/move", methods=["POST"])
@jwt_required()
def bulk_move_files():
    username = get_jwt_identity()
    data = request.get_json()
    filepaths = data.get('filepaths', [])
    destination = data.get('destination', '').strip()

    if not filepaths:
        return jsonify({"msg": "No files specified"}), 400

    destination = validate_path(username, destination)
    user_dir = STORAGE_DIR / username
    dest_dir = user_dir / destination if destination else user_dir

    if not dest_dir.exists():
        return jsonify({"msg": "Destination folder not found"}), 404

    moved = []
    errors = []

    for filepath in filepaths:
        filepath = validate_path(username, filepath)
        file_path = user_dir / filepath

        try:
            file_path.resolve().relative_to(user_dir.resolve())
            if file_path.exists() and file_path.is_file():
                dest_path = dest_dir / file_path.name

                if dest_path.exists():
                    name, ext = os.path.splitext(file_path.name)
                    counter = 1
                    while dest_path.exists():
                        dest_path = dest_dir / f"{name}_{counter}{ext}"
                        counter += 1

                shutil.move(str(file_path), str(dest_path))
                moved.append(filepath)
                logger.info(f"Bulk move: {username}/{filepath} -> {destination}")
            else:
                errors.append(filepath)
        except Exception as e:
            logger.error(f"Bulk move error for {username}/{filepath}: {str(e)}")
            errors.append(filepath)

    msg = f"Moved {len(moved)} file(s)"
    if errors:
        msg += f". {len(errors)} failed"

    return jsonify({"msg": msg, "moved": moved, "errors": errors}), 200


@app.route("/api/bulk/share", methods=["POST"])
@jwt_required()
def bulk_share_files():
    username = get_jwt_identity()
    user = user_manager.get_user(username)
    data = request.get_json()
    filepaths = data.get('filepaths', [])

    if not filepaths:
        return jsonify({"msg": "No files specified"}), 400

    user_dir = STORAGE_DIR / username
    shared = []
    errors = []

    for filepath in filepaths:
        filepath = validate_path(username, filepath)
        file_path = user_dir / filepath

        try:
            file_path.resolve().relative_to(user_dir.resolve())
            if file_path.exists() and file_path.is_file():
                stat = file_path.stat()
                file_id = shared_manager.request_share(
                    username,
                    filepath,
                    file_path.name,
                    stat.st_size,
                    get_file_type(file_path.name)
                )

                if should_auto_approve(user):
                    shared_manager.approve_share(file_id)

                shared.append(filepath)
                logger.info(f"Bulk share: {username}/{filepath}")
            else:
                errors.append(filepath)
        except Exception as e:
            logger.error(f"Bulk share error for {username}/{filepath}: {str(e)}")
            errors.append(filepath)

    msg = f"Shared {len(shared)} file(s)"
    if errors:
        msg += f". {len(errors)} failed"

    return jsonify({"msg": msg, "shared": shared, "errors": errors}), 200


@app.route("/api/preview/docx/<path:filepath>", methods=["GET"])
def preview_docx(filepath):
    token = request.args.get('token')
    if not token:
        return jsonify({"msg": "No token provided"}), 401

    try:
        from flask_jwt_extended import decode_token
        decoded = decode_token(token)
        username = decoded['sub']
    except Exception:
        return jsonify({"msg": "Invalid token"}), 401

    filepath = validate_path(username, filepath)
    user_dir = STORAGE_DIR / username
    file_path = user_dir / filepath

    try:
        file_path.resolve().relative_to(user_dir.resolve())
    except ValueError:
        return jsonify({"msg": "Access denied"}), 403

    if not file_path.exists():
        return jsonify({"msg": "File not found"}), 404

    try:
        import mammoth
        with open(file_path, "rb") as docx_file:
            result = mammoth.convert_to_html(docx_file)
            html = result.value

        html_output = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                body {{
                    font-family: 'Calibri', 'Arial', sans-serif;
                    padding: 40px;
                    max-width: 800px;
                    margin: 0 auto;
                    background: #fff;
                    color: #000;
                }}
                img {{ max-width: 100%; height: auto; }}
            </style>
        </head>
        <body>
            {html}
        </body>
        </html>
        """
        return html_output, 200, {'Content-Type': 'text/html; charset=utf-8'}
    except Exception as e:
        logger.error(f"DOCX preview error: {str(e)}")
        return jsonify({"msg": "Failed to preview DOCX file"}), 500


@app.route("/api/preview/xlsx/<path:filepath>", methods=["GET"])
def preview_xlsx(filepath):
    token = request.args.get('token')
    if not token:
        return jsonify({"msg": "No token provided"}), 401

    try:
        from flask_jwt_extended import decode_token
        decoded = decode_token(token)
        username = decoded['sub']
    except Exception:
        return jsonify({"msg": "Invalid token"}), 401

    filepath = validate_path(username, filepath)
    user_dir = STORAGE_DIR / username
    file_path = user_dir / filepath

    try:
        file_path.resolve().relative_to(user_dir.resolve())
    except ValueError:
        return jsonify({"msg": "Access denied"}), 403

    if not file_path.exists():
        return jsonify({"msg": "File not found"}), 404

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)

        html_parts = ['<!DOCTYPE html><html><head><meta charset="UTF-8"><style>']
        html_parts.append("""
            body { font-family: 'Calibri', 'Arial', sans-serif; padding: 20px; background: #f5f5f5; }
            .sheet-tabs { margin-bottom: 20px; border-bottom: 2px solid #ddd; }
            .sheet-tab { display: inline-block; padding: 10px 20px; margin-right: 5px; background: #fff;
                border: 1px solid #ddd; border-bottom: none; cursor: pointer; border-radius: 5px 5px 0 0; }
            .sheet-tab.active { background: #4CAF50; color: white; font-weight: bold; }
            .sheet-content { display: none; background: white; padding: 20px; border-radius: 5px;
                box-shadow: 0 2px 4px rgba(0,0,0,0.1); overflow-x: auto; }
            .sheet-content.active { display: block; }
            table { border-collapse: collapse; width: 100%; background: white; }
            th, td { border: 1px solid #ddd; padding: 8px; text-align: left; white-space: nowrap; }
            th { background-color: #4CAF50; color: white; font-weight: bold; }
            tr:nth-child(even) { background-color: #f9f9f9; }
            tr:hover { background-color: #f5f5f5; }
        </style>
        <script>
            function showSheet(idx) {
                document.querySelectorAll('.sheet-content').forEach(s => s.classList.remove('active'));
                document.querySelectorAll('.sheet-tab').forEach(t => t.classList.remove('active'));
                document.getElementById('sheet-' + idx).classList.add('active');
                document.getElementById('tab-' + idx).classList.add('active');
            }
        </script>
        </head><body>
        """)

        html_parts.append('<div class="sheet-tabs">')
        for idx, sheet_name in enumerate(wb.sheetnames):
            active_class = 'active' if idx == 0 else ''
            html_parts.append(f'<div class="sheet-tab {active_class}" id="tab-{idx}" onclick="showSheet({idx})">{sheet_name}</div>')
        html_parts.append('</div>')

        for idx, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]
            active_class = 'active' if idx == 0 else ''
            html_parts.append(f'<div class="sheet-content {active_class}" id="sheet-{idx}">')
            html_parts.append(f'<h2>{sheet_name}</h2><table>')

            for row_idx, row in enumerate(ws.iter_rows(max_row=min(100, ws.max_row)), 1):
                html_parts.append('<tr>')
                for cell in row:
                    value = cell.value if cell.value is not None else ''
                    tag = 'th' if row_idx == 1 else 'td'
                    html_parts.append(f'<{tag}>{value}</{tag}>')
                html_parts.append('</tr>')

            if ws.max_row > 100:
                html_parts.append(f'<tr><td colspan="{ws.max_column}"><em>... showing first 100 rows of {ws.max_row}</em></td></tr>')

            html_parts.append('</table></div>')

        html_parts.append('</body></html>')
        return ''.join(html_parts), 200, {'Content-Type': 'text/html; charset=utf-8'}
    except Exception as e:
        logger.error(f"XLSX preview error: {str(e)}")
        return jsonify({"msg": "Failed to preview Excel file"}), 500


@app.route("/api/users", methods=["GET"])
@jwt_required()
def get_users():
    username = get_jwt_identity()
    user = user_manager.get_user(username)

    if user["role"] != "admin":
        return jsonify({"msg": "Admin access required"}), 403

    users = user_manager.list_users()
    return jsonify({"users": users}), 200


@app.route("/api/users/pending", methods=["GET"])
@jwt_required()
def get_pending_users():
    username = get_jwt_identity()
    user = user_manager.get_user(username)

    if user["role"] != "admin":
        return jsonify({"msg": "Admin access required"}), 403

    users = user_manager.list_users(status="pending")
    return jsonify({"users": users}), 200


@app.route("/api/users", methods=["POST"])
@jwt_required()
def add_user():
    username = get_jwt_identity()
    user = user_manager.get_user(username)

    if user["role"] != "admin":
        return jsonify({"msg": "Admin access required"}), 403

    data = request.get_json()
    new_username = data.get("username", "").strip()
    password = data.get("password", "")
    role = data.get("role", "user")

    if role not in ["user", "admin"]:
        return jsonify({"msg": "Invalid role"}), 400

    success, msg = user_manager.add_user(new_username, password, role)

    if success:
        logger.info(f"User created by {username}: {new_username} ({role})")
        return jsonify({"msg": msg}), 200
    return jsonify({"msg": msg}), 400


@app.route("/api/users/<target_username>/approve", methods=["POST"])
@jwt_required()
def approve_user(target_username):
    username = get_jwt_identity()
    user = user_manager.get_user(username)

    if user["role"] != "admin":
        return jsonify({"msg": "Admin access required"}), 403

    success, msg = user_manager.approve_user(target_username)

    if success:
        logger.info(f"User approved by {username}: {target_username}")
        return jsonify({"msg": msg}), 200
    return jsonify({"msg": msg}), 400


@app.route("/api/users/<target_username>/reject", methods=["POST"])
@jwt_required()
def reject_user(target_username):
    username = get_jwt_identity()
    user = user_manager.get_user(username)

    if user["role"] != "admin":
        return jsonify({"msg": "Admin access required"}), 403

    success, msg = user_manager.reject_user(target_username)

    if success:
        logger.info(f"User rejected by {username}: {target_username}")
        return jsonify({"msg": msg}), 200
    return jsonify({"msg": msg}), 400


@app.route("/api/users/<target_username>", methods=["DELETE"])
@jwt_required()
def delete_user(target_username):
    username = get_jwt_identity()
    user = user_manager.get_user(username)

    if user["role"] != "admin":
        return jsonify({"msg": "Admin access required"}), 403

    success, msg = user_manager.delete_user(target_username)

    if success:
        logger.info(f"User deleted by {username}: {target_username}")
        return jsonify({"msg": msg}), 200
    return jsonify({"msg": msg}), 400


@app.route("/api/users/<target_username>/trusted", methods=["POST"])
@jwt_required()
def set_trusted_uploader(target_username):
    """Toggle trusted uploader status — admin only."""
    username = get_jwt_identity()
    user = user_manager.get_user(username)

    if user["role"] != "admin":
        return jsonify({"msg": "Admin access required"}), 403

    data = request.get_json()
    value = bool(data.get("trusted_uploader", False))

    success, msg = user_manager.set_trusted_uploader(target_username, value)
    if success:
        logger.info(f"Trusted uploader set to {value} for {target_username} by {username}")
        return jsonify({"msg": msg}), 200
    return jsonify({"msg": msg}), 400


@app.route("/api/users/<target_username>/auto_share", methods=["POST"])
@jwt_required()
def set_auto_share(target_username):
    """Toggle auto-share status — admin only."""
    username = get_jwt_identity()
    user = user_manager.get_user(username)

    if user["role"] != "admin":
        return jsonify({"msg": "Admin access required"}), 403

    data = request.get_json()
    value = bool(data.get("auto_share", False))

    success, msg = user_manager.set_auto_share(target_username, value)
    if success:
        logger.info(f"Auto-share set to {value} for {target_username} by {username}")
        return jsonify({"msg": msg}), 200
    return jsonify({"msg": msg}), 400


@app.route("/api/account/change-password", methods=["POST"])
@jwt_required()
def change_password():
    username = get_jwt_identity()
    data = request.get_json()
    current_password = data.get("current_password", "")
    new_password = data.get("new_password", "")

    if not current_password or not new_password:
        return jsonify({"msg": "All fields required"}), 400

    if not user_manager.authenticate(username, current_password):
        return jsonify({"msg": "Current password incorrect"}), 401

    if len(new_password) < 6:
        return jsonify({"msg": "New password must be at least 6 characters"}), 400

    user_manager.users[username]["password_hash"] = bcrypt.hashpw(new_password.encode(), bcrypt.gensalt()).decode()
    user_manager.save_users()

    logger.info(f"Password changed for user: {username}")
    return jsonify({"msg": "Password changed successfully"}), 200


@app.route("/api/account/change-username", methods=["POST"])
@jwt_required()
def change_username():
    old_username = get_jwt_identity()
    data = request.get_json()
    new_username = data.get("new_username", "").strip()
    password = data.get("password", "")

    if not new_username or not password:
        return jsonify({"msg": "All fields required"}), 400

    if not validate_username(new_username):
        return jsonify({"msg": "Invalid username format"}), 400

    if new_username in user_manager.users:
        return jsonify({"msg": "Username already exists"}), 400

    if not user_manager.authenticate(old_username, password):
        return jsonify({"msg": "Password incorrect"}), 401

    user_data = user_manager.users[old_username]
    del user_manager.users[old_username]
    user_manager.users[new_username] = user_data
    user_manager.save_users()

    old_dir = STORAGE_DIR / old_username
    new_dir = STORAGE_DIR / new_username
    if old_dir.exists():
        old_dir.rename(new_dir)

    for entry in shared_manager.shared_files.get("pending", []):
        if entry["username"] == old_username:
            entry["username"] = new_username

    for entry in shared_manager.shared_files.get("approved", []):
        if entry["username"] == old_username:
            entry["username"] = new_username

    for entry in shared_manager.shared_files.get("pending_folders", []):
        if entry["username"] == old_username:
            entry["username"] = new_username

    for entry in shared_manager.shared_files.get("folders", []):
        if entry["username"] == old_username:
            entry["username"] = new_username

    shared_manager.save_shared_files()

    new_token = create_access_token(identity=new_username)

    logger.info(f"Username changed from {old_username} to {new_username}")
    return jsonify({
        "msg": "Username changed successfully",
        "access_token": new_token,
        "new_username": new_username
    }), 200


@app.route("/api/upload/chunk", methods=["POST"])
@jwt_required()
def upload_chunk():
    username = get_jwt_identity()
    user = user_manager.get_user(username)

    upload_id    = request.form.get("upload_id", "").strip()
    filename     = request.form.get("filename", "").strip()
    chunk_index  = int(request.form.get("chunk_index", 0))
    total_chunks = int(request.form.get("total_chunks", 1))
    folder_path  = validate_path(username, request.form.get("folder", "").strip())

    if not upload_id or not filename:
        return jsonify({"msg": "Missing upload_id or filename"}), 400

    if not allowed_file(filename):
        return jsonify({"msg": f"File type not allowed: {filename}"}), 400

    chunk_file = request.files.get("file")
    if not chunk_file:
        return jsonify({"msg": "No chunk data"}), 400

    tmp_dir = BASE_DIR / "tmp_uploads" / upload_id
    tmp_dir.mkdir(parents=True, exist_ok=True)

    chunk_path = tmp_dir / f"chunk_{chunk_index:06d}"
    chunk_file.save(str(chunk_path))

    received = len(list(tmp_dir.glob("chunk_*")))
    if received < total_chunks:
        return jsonify({
            "msg": f"Chunk {chunk_index + 1}/{total_chunks} received",
            "done": False
        }), 200

    user_dir = STORAGE_DIR / username
    base_dir  = user_dir / folder_path if folder_path else user_dir
    is_folder = request.form.get("is_folder_upload", "false") == "true"

    if is_folder and ("/" in filename or "\\" in filename):
        rel_path    = sanitize_relative_path(filename)
        if not rel_path:
            rel_path = secure_filename(filename.replace("\\", "/").rsplit("/", 1)[-1]) or "file"
        target_path = base_dir / rel_path
    else:
        safe_name = secure_filename(filename)
        if not safe_name:
            safe_name = re.sub(r'[^\w\-.]', '_', filename).strip('._') or 'file'
        target_path = base_dir / safe_name

    target_path.parent.mkdir(parents=True, exist_ok=True)

    if target_path.exists():
        name, ext = os.path.splitext(target_path.name)
        counter = 1
        while target_path.exists():
            target_path = target_path.parent / f"{name}_{counter}{ext}"
            counter += 1

    # Check if uploading into an already-shared folder
    shared_folder_entry = shared_manager.get_shared_folder_for_path(username, folder_path) if folder_path else None

    try:
        with open(target_path, "wb") as out:
            for i in range(total_chunks):
                cp = tmp_dir / f"chunk_{i:06d}"
                with open(cp, "rb") as c:
                    out.write(c.read())

        shutil.rmtree(tmp_dir, ignore_errors=True)

        rel = str(target_path.relative_to(user_dir))

        # Auto-share if user has auto_share flag OR uploading into a shared folder
        if user.get("auto_share", False) or shared_folder_entry:
            stat = target_path.stat()
            fid  = shared_manager.request_share(
                username, rel, target_path.name,
                stat.st_size, get_file_type(target_path.name)
            )
            shared_manager.approve_share(fid)

        user_manager.update_storage_used(username)
        logger.info(f"Chunked upload complete: {username}/{rel}")

        return jsonify({
            "msg": f"'{target_path.name}' uploaded successfully",
            "done": True,
            "filepath": rel
        }), 200

    except Exception as e:
        shutil.rmtree(tmp_dir, ignore_errors=True)
        logger.error(f"Chunk reassembly error for {username}/{filename}: {e}")
        return jsonify({"msg": f"Upload failed: {str(e)}"}), 500


@app.route("/api/upload/chunk/cancel", methods=["POST"])
@jwt_required()
def cancel_chunked_upload():
    data = request.get_json()
    upload_id = (data or {}).get("upload_id", "").strip()
    if upload_id:
        tmp_dir = BASE_DIR / "tmp_uploads" / upload_id
        shutil.rmtree(tmp_dir, ignore_errors=True)
    return jsonify({"msg": "Upload cancelled"}), 200


# ==================== TEXT FILE EDITOR ENDPOINTS ====================

EDITABLE_EXTENSIONS = {'md', 'txt', 'json', 'xml', 'csv', 'log', 'yaml', 'yml',
                        'toml', 'ini', 'cfg', 'conf'}
MAX_EDITABLE_SIZE = 5 * 1024 * 1024  # 5 MB


def is_editable(filename):
    if '.' not in filename:
        return False
    ext = filename.rsplit('.', 1)[1].lower()
    return ext in EDITABLE_EXTENSIONS


@app.route("/api/file/read/<path:filepath>", methods=["GET"])
@jwt_required()
def read_file_content(filepath):
    """Read a text file's content for editing."""
    username = get_jwt_identity()
    filepath = validate_path(username, filepath)
    user_dir = STORAGE_DIR / username
    file_path = user_dir / filepath

    try:
        file_path.resolve().relative_to(user_dir.resolve())
    except ValueError:
        return jsonify({"msg": "Access denied"}), 403

    if not file_path.exists():
        return jsonify({"msg": "File not found"}), 404

    if not is_editable(file_path.name):
        return jsonify({"msg": "File type not editable"}), 400

    if file_path.stat().st_size > MAX_EDITABLE_SIZE:
        return jsonify({"msg": "File too large to edit (max 5 MB)"}), 400

    try:
        content = file_path.read_text(encoding='utf-8')
        return jsonify({
            "content": content,
            "filename": file_path.name,
            "filepath": filepath,
            "size": file_path.stat().st_size
        }), 200
    except UnicodeDecodeError:
        return jsonify({"msg": "File is not valid UTF-8 text"}), 400
    except Exception as e:
        logger.error(f"Read file error: {e}")
        return jsonify({"msg": f"Failed to read file: {str(e)}"}), 500


@app.route("/api/file/write/<path:filepath>", methods=["POST"])
@jwt_required()
def write_file_content(filepath):
    """Write/save content to a text file."""
    username = get_jwt_identity()
    filepath = validate_path(username, filepath)
    user_dir = STORAGE_DIR / username
    file_path = user_dir / filepath

    try:
        file_path.resolve().relative_to(user_dir.resolve())
    except ValueError:
        return jsonify({"msg": "Access denied"}), 403

    if not is_editable(file_path.name):
        return jsonify({"msg": "File type not editable"}), 400

    data = request.get_json()
    content = data.get("content", "")

    if len(content.encode('utf-8')) > MAX_EDITABLE_SIZE:
        return jsonify({"msg": "Content too large (max 5 MB)"}), 400

    try:
        file_path.parent.mkdir(parents=True, exist_ok=True)
        file_path.write_text(content, encoding='utf-8')
        user_manager.update_storage_used(username)
        logger.info(f"File written: {username}/{filepath}")
        return jsonify({"msg": "Saved", "size": file_path.stat().st_size}), 200
    except Exception as e:
        logger.error(f"Write file error: {e}")
        return jsonify({"msg": f"Failed to save: {str(e)}"}), 500


@app.route("/api/file/create", methods=["POST"])
@jwt_required()
def create_text_file():
    """Create a new text/markdown file."""
    username = get_jwt_identity()
    data = request.get_json()
    folder_path = validate_path(username, data.get("folder", "").strip())
    filename = data.get("filename", "").strip()

    if not filename:
        return jsonify({"msg": "Filename required"}), 400

    # Ensure .md extension by default if no extension given
    if '.' not in filename:
        filename += '.md'

    ext = filename.rsplit('.', 1)[1].lower()
    if ext not in EDITABLE_EXTENSIONS:
        return jsonify({"msg": f"Extension .{ext} not allowed for text files"}), 400

    # Sanitize filename, preserving leading dot for hidden files
    leading_dot = filename.startswith('.')
    inner = filename.lstrip('.')
    safe = secure_filename(inner) if inner else ''
    if not safe:
        return jsonify({"msg": "Invalid filename"}), 400
    filename = ('.' + safe) if leading_dot else safe

    user_dir = STORAGE_DIR / username
    base = user_dir / folder_path if folder_path else user_dir
    file_path = base / filename

    try:
        file_path.resolve().relative_to(user_dir.resolve())
    except ValueError:
        return jsonify({"msg": "Access denied"}), 403

    if file_path.exists():
        return jsonify({"msg": "File already exists"}), 400

    try:
        base.mkdir(parents=True, exist_ok=True)
        initial = f"# {filename.rsplit('.', 1)[0]}\n\n" if filename.endswith('.md') else ""
        file_path.write_text(initial, encoding='utf-8')
        user_manager.update_storage_used(username)
        rel = str(file_path.relative_to(user_dir))
        logger.info(f"Text file created: {username}/{rel}")
        return jsonify({"msg": "File created", "filepath": rel, "filename": filename}), 200
    except Exception as e:
        logger.error(f"Create file error: {e}")
        return jsonify({"msg": f"Failed to create file: {str(e)}"}), 500



# ──────────────────────────────────────────────────────────────────
#  PINNED FOLDERS  (server-synced per user)
# ──────────────────────────────────────────────────────────────────

@app.route("/api/pinned", methods=["GET"])
@jwt_required()
def get_pinned():
    username = get_jwt_identity()
    user = user_manager.get_user(username)
    pinned = user.get("pinned_folders", [])
    return jsonify({"pinned": pinned}), 200


@app.route("/api/pinned", methods=["POST"])
@jwt_required()
def add_pinned():
    username = get_jwt_identity()
    data = request.get_json() or {}
    path = data.get("path", "").strip()
    name = data.get("name", path.split("/")[-1] if path else "Root").strip()
    if not path:
        return jsonify({"msg": "Path required"}), 400
    user = user_manager.get_user(username)
    pinned = user.get("pinned_folders", [])
    if any(p["path"] == path for p in pinned):
        return jsonify({"msg": "Already pinned"}), 409
    pinned.append({"path": path, "name": name})
    user_manager.users[username]["pinned_folders"] = pinned
    user_manager.save_users()
    return jsonify({"msg": "Pinned", "pinned": pinned}), 200


@app.route("/api/pinned", methods=["DELETE"])
@jwt_required()
def remove_pinned():
    username = get_jwt_identity()
    data = request.get_json() or {}
    path = data.get("path", "").strip()
    user = user_manager.get_user(username)
    pinned = [p for p in user.get("pinned_folders", []) if p["path"] != path]
    user_manager.users[username]["pinned_folders"] = pinned
    user_manager.save_users()
    return jsonify({"msg": "Unpinned", "pinned": pinned}), 200


# ──────────────────────────────────────────────────────────────────
#  FAVORITES  (server-synced per user)
# ──────────────────────────────────────────────────────────────────

@app.route("/api/favorites", methods=["GET"])
@jwt_required()
def get_favorites():
    username = get_jwt_identity()
    user = user_manager.get_user(username)
    favs = user.get("favorites", [])
    return jsonify({"favorites": favs}), 200


@app.route("/api/favorites", methods=["POST"])
@jwt_required()
def add_favorite():
    username = get_jwt_identity()
    data = request.get_json() or {}
    path = data.get("path", "").strip()
    name = data.get("name", path.split("/")[-1] if path else "").strip()
    if not path:
        return jsonify({"msg": "Path required"}), 400
    user = user_manager.get_user(username)
    favs = user.get("favorites", [])
    if any(f["path"] == path for f in favs):
        return jsonify({"msg": "Already favorited"}), 409
    favs.append({"path": path, "name": name})
    user_manager.users[username]["favorites"] = favs
    user_manager.save_users()
    return jsonify({"msg": "Favorited", "favorites": favs}), 200


@app.route("/api/favorites", methods=["DELETE"])
@jwt_required()
def remove_favorite():
    username = get_jwt_identity()
    data = request.get_json() or {}
    path = data.get("path", "").strip()
    user = user_manager.get_user(username)
    favs = [f for f in user.get("favorites", []) if f["path"] != path]
    user_manager.users[username]["favorites"] = favs
    user_manager.save_users()
    return jsonify({"msg": "Unfavorited", "favorites": favs}), 200


# ──────────────────────────────────────────────────────────────────
#  RENAME FILE/FOLDER
# ──────────────────────────────────────────────────────────────────

@app.route("/api/rename", methods=["POST"])
@jwt_required()
def rename_item():
    username = get_jwt_identity()
    data = request.get_json() or {}
    filepath = validate_path(username, (data.get('filepath') or '').strip())
    new_name = (data.get('new_name') or '').strip()

    if not filepath or not new_name:
        return jsonify({"msg": "filepath and new_name required"}), 400

    user_dir = STORAGE_DIR / username
    src_path = user_dir / filepath

    try:
        src_path.resolve().relative_to(user_dir.resolve())
    except ValueError:
        return jsonify({"msg": "Access denied"}), 403

    if not src_path.exists():
        return jsonify({"msg": "Not found"}), 404

    parent = src_path.parent
    # Sanitize new name
    if src_path.is_file():
        ext = src_path.suffix
        name_no_ext = new_name if new_name.endswith(ext) else new_name
        safe = secure_filename(name_no_ext)
        if not safe:
            return jsonify({"msg": "Invalid name"}), 400
        dst_path = parent / safe
    else:
        safe = secure_filename(new_name)
        if not safe:
            return jsonify({"msg": "Invalid name"}), 400
        dst_path = parent / safe

    if dst_path.exists():
        return jsonify({"msg": "A file with that name already exists"}), 409

    try:
        src_path.rename(dst_path)
        user_manager.update_storage_used(username)
        return jsonify({"msg": "Renamed", "new_path": str(dst_path.relative_to(user_dir))}), 200
    except Exception as e:
        return jsonify({"msg": str(e)}), 500


# ──────────────────────────────────────────────────────────────────
#  MUSIC PLAYER  (shared /tracks directory, JWT auth, no passwords)
# ──────────────────────────────────────────────────────────────────

try:
    import yt_dlp as _yt_dlp
    YT_DLP_AVAILABLE = True
except ImportError:
    YT_DLP_AVAILABLE = False

import threading as _threading

MUSIC_DIR = BASE_DIR / "tracks"
MUSIC_DIR.mkdir(exist_ok=True)
MUSIC_EXTENSIONS = {'mp3', 'flac', 'wav', 'ogg', 'm4a', 'aac', 'opus', 'wma', 'webm'}
_dl_status = {}   # download_id -> {status, name, error, filename}


def _safe_music_path(filename):
    path = (MUSIC_DIR / filename).resolve()
    try:
        path.relative_to(MUSIC_DIR.resolve())
        return path
    except ValueError:
        return None


def _music_size_fmt(b):
    for u in ['B','KB','MB','GB']:
        if b < 1024: return f"{b:.1f} {u}"
        b /= 1024
    return f"{b:.1f} TB"


@app.route("/api/music/tracks", methods=["GET"])
@jwt_required()
def music_list_tracks():
    try:
        files = []
        for f in MUSIC_DIR.iterdir():
            if f.is_file() and f.suffix.lstrip('.').lower() in MUSIC_EXTENSIONS:
                st = f.stat()
                files.append({"name": f.name, "size": st.st_size,
                               "size_fmt": _music_size_fmt(st.st_size),
                               "modified": st.st_mtime})
        files.sort(key=lambda x: x["name"].lower())
        return jsonify({"tracks": files}), 200
    except Exception as e:
        return jsonify({"msg": str(e)}), 500


@app.route("/api/music/stream/<path:filename>")
@jwt_required()
def music_stream(filename):
    path = _safe_music_path(filename)
    if not path or not path.exists():
        return abort(404)
    mime = mimetypes.guess_type(str(path))[0] or "audio/mpeg"
    # Use send_file with conditional=True — Flask handles Range requests,
    # ETags and Last-Modified automatically → instant seek, zero delay
    response = make_response(send_file(str(path), mimetype=mime, conditional=True))
    response.headers["Accept-Ranges"] = "bytes"
    return response


@app.route("/api/music/upload", methods=["POST"])
@jwt_required()
def music_upload():
    if "file" not in request.files:
        return jsonify({"msg": "No file"}), 400
    file = request.files["file"]
    custom_name = request.form.get("name", "").strip()
    _, ext = os.path.splitext(file.filename or "")
    ext = ext.lower() or ".mp3"
    if ext.lstrip('.') not in MUSIC_EXTENSIONS:
        return jsonify({"msg": f"Unsupported format: {ext}"}), 400
    base = custom_name if custom_name else os.path.splitext(file.filename or "track")[0]
    safe = re.sub(r'[/\\:*?"<>|]', '_', base).strip() or "track"
    fname = safe + ext
    fpath = MUSIC_DIR / fname
    c = 1
    while fpath.exists():
        fpath = MUSIC_DIR / f"{safe}_{c}{ext}"; c += 1
    file.save(fpath)
    logger.info(f"Music upload: {fname}")
    return jsonify({"success": True, "filename": fpath.name}), 201


@app.route("/api/music/youtube", methods=["POST"])
@jwt_required()
def music_youtube():
    if not YT_DLP_AVAILABLE:
        return jsonify({"msg": "yt-dlp not installed on server. Run: pip install yt-dlp"}), 500
    data = request.get_json() or {}
    url    = data.get("url", "").strip()
    name   = data.get("name", "").strip()
    fmt    = data.get("format", "mp3").strip().lower()
    thumb  = bool(data.get("thumbnail", True))
    if fmt not in ("mp3", "wav"):
        fmt = "mp3"
    if not url:
        return jsonify({"msg": "URL required"}), 400
    did = secrets.token_urlsafe(8)
    _dl_status[did] = {"status": "downloading", "name": name or url, "error": None, "filename": None, "count": 0}

    def _lookup_artist_musicbrainz(title, duration_sec=None):
        """Try MusicBrainz to find the actual artist for a track title."""
        try:
            q = _urllib_parse.urlencode({"query": f'recording:"{title}"', "fmt": "json", "limit": "3"})
            req = _urllib_req.Request(
                f"https://musicbrainz.org/ws/2/recording?{q}",
                headers={"User-Agent": "NAS-MusicPlayer/1.0 (nas@localhost)"}
            )
            with _urllib_req.urlopen(req, timeout=5) as r:
                mb = json.loads(r.read())
            recordings = mb.get("recordings", [])
            if not recordings:
                return None
            # If we have duration, pick the closest match
            best = recordings[0]
            if duration_sec:
                for rec in recordings:
                    rec_dur = rec.get("length", 0) / 1000
                    if abs(rec_dur - duration_sec) < 10:
                        best = rec
                        break
            artists = best.get("artist-credit", [])
            if artists:
                return artists[0].get("artist", {}).get("name")
        except Exception:
            pass
        return None

    def _build_filename(info, custom_name):
        """Build a clean 'Artist - Title' filename from yt-dlp info dict."""
        if custom_name:
            return re.sub(r'[/\\:*?"<>|]', '_', custom_name).strip()
        title   = (info.get("title") or "Unknown").strip()
        # yt-dlp exposes artist from video description / music metadata
        artist  = (info.get("artist") or info.get("creator") or "").strip()
        # Fallback: try MusicBrainz lookup
        if not artist or artist.lower() in ("various artists", ""):
            duration = info.get("duration")
            artist = _lookup_artist_musicbrainz(title, duration) or artist
        # Last resort: use channel/uploader name (e.g. "Coke Studio", "T-Series")
        if not artist:
            artist = (info.get("uploader") or info.get("channel") or "").strip()
            # Strip trailing " - Topic" suffix YouTube adds for auto-generated channels
            artist = re.sub(r'\s*-\s*Topic\s*$', '', artist, flags=re.IGNORECASE).strip()
        # Clean title — strip "official video", "(lyrics)", etc.
        clean_title = re.sub(
            r'\s*[\(\[].*?(official|video|audio|lyrics|hd|4k|mv|music|feat\.?|ft\.?).*?[\)\]]',
            '', title, flags=re.IGNORECASE
        ).strip()
        if artist:
            return re.sub(r'[/\\:*?"<>|]', '_', f"{artist} - {clean_title}").strip()
        return re.sub(r'[/\\:*?"<>|]', '_', clean_title).strip()

    def run():
        # First extract info without downloading to get metadata
        probe_opts = {"quiet": True, "no_warnings": True, "noplaylist": True,
                      "skip_download": True, "writeinfojson": False}
        try:
            with _yt_dlp.YoutubeDL(probe_opts) as ydl:
                info = ydl.extract_info(url, download=False)
        except Exception as e:
            _dl_status[did].update({"status": "error", "error": str(e)})
            return

        fname_base = _build_filename(info, name)
        tmpl = str(MUSIC_DIR / (fname_base + ".%(ext)s"))

        postproc = [{"key": "FFmpegExtractAudio", "preferredcodec": fmt,
                     "preferredquality": "192" if fmt == "mp3" else "0"}]
        if thumb:
            postproc += [{"key": "FFmpegMetadata"}]
            if fmt == "mp3":
                postproc += [{"key": "EmbedThumbnail"}]
        opts = {
            "format": "bestaudio/best", "outtmpl": tmpl,
            "postprocessors": postproc,
            "writethumbnail": thumb,
            "noplaylist": True, "quiet": True, "no_warnings": True,
        }
        try:
            with _yt_dlp.YoutubeDL(opts) as ydl:
                ydl.download([url])
            display = fname_base
            _dl_status[did].update({"status": "done", "filename": display})
            logger.info(f"YT download done: {url} -> {display}")
        except Exception as e:
            _dl_status[did].update({"status": "error", "error": str(e)})
            logger.error(f"YT download error: {e}")

    _threading.Thread(target=run, daemon=True).start()
    return jsonify({"success": True, "download_id": did}), 202


@app.route("/api/music/channel", methods=["POST"])
@jwt_required()
def music_channel():
    """Download all audio from a YouTube channel or playlist URL, then auto-create a playlist."""
    if not YT_DLP_AVAILABLE:
        return jsonify({"msg": "yt-dlp not installed on server. Run: pip install yt-dlp"}), 500
    data       = request.get_json() or {}
    url        = data.get("url", "").strip()
    fmt        = data.get("format", "mp3").strip().lower()
    thumb      = bool(data.get("thumbnail", True))
    limit      = int(data.get("limit", 0))
    pl_name    = data.get("playlist_name", "").strip()
    username   = get_jwt_identity()
    if fmt not in ("mp3", "wav"):
        fmt = "mp3"
    if not url:
        return jsonify({"msg": "URL required"}), 400
    # Support artist Releases tab: youtube.com/@Artist/releases → append /videos or keep as-is
    # yt-dlp handles /releases tab natively; just pass through
    # Normalize shorts channel URLs: /@artist → /@artist/videos for full discography if no tab specified
    import urllib.parse as _up
    _parsed = _up.urlparse(url)
    _path = _parsed.path.rstrip('/')
    # If it ends in /releases, /singles, /albums — keep as-is (yt-dlp understands them)
    # If bare channel URL with no tab, add /videos to get all uploads
    _tab_suffixes = ('/videos', '/releases', '/shorts', '/streams', '/playlists')
    if (_path.startswith('/@') or '/channel/' in _path or '/c/' in _path) and \
       not any(_path.endswith(s) for s in _tab_suffixes) and \
       'list=' not in url:
        url = url.rstrip('/') + '/videos'
    did = secrets.token_urlsafe(8)
    _dl_status[did] = {"status": "downloading", "name": url, "error": None,
                        "filename": None, "count": 0, "total": None, "playlist": None}

    downloaded_files = []

    def progress_hook(d):
        if d.get("status") == "finished":
            fpath = d.get("filename", "")
            if fpath:
                downloaded_files.append(os.path.basename(fpath))
            _dl_status[did]["count"] = len(downloaded_files)

    def run():
        tmpl = str(MUSIC_DIR / '%(artist)s - %(title)s.%(ext)s')
        # Fallback template if no artist tag
        postproc = [{"key": "FFmpegExtractAudio", "preferredcodec": fmt,
                     "preferredquality": "192" if fmt == "mp3" else "0"}]
        if thumb:
            postproc += [{"key": "FFmpegMetadata"}]
            if fmt == "mp3":
                postproc += [{"key": "EmbedThumbnail"}]
        opts = {
            "format": "bestaudio/best",
            "outtmpl": tmpl,
            "postprocessors": postproc,
            "writethumbnail": thumb,
            "ignoreerrors": True,
            "noplaylist": False,
            "quiet": True, "no_warnings": True,
            "progress_hooks": [progress_hook],
        }
        if limit > 0:
            opts["playlistend"] = limit
        try:
            with _yt_dlp.YoutubeDL(opts) as ydl:
                info = ydl.extract_info(url, download=True)

            # Resolve playlist name
            auto_pl_name = pl_name
            if not auto_pl_name and info:
                auto_pl_name = (info.get("title") or info.get("uploader") or
                                info.get("channel") or "Downloaded Playlist")

            # Normalise filenames — FFmpegExtractAudio changes extension
            ext = "." + fmt
            final_tracks = []
            for raw in downloaded_files:
                # Strip old extension, add new one
                base = re.sub(r'\.[^.]+$', '', raw)
                candidate = base + ext
                if (MUSIC_DIR / candidate).exists():
                    final_tracks.append(candidate)
                elif (MUSIC_DIR / raw).exists():
                    final_tracks.append(raw)

            # Auto-create or overwrite playlist in user prefs
            if final_tracks and auto_pl_name:
                user = user_manager.get_user(username)
                playlists = user.get("music_playlists", [])
                # Find existing playlist with same name → overwrite
                existing = next((p for p in playlists if p["name"] == auto_pl_name), None)
                if existing:
                    existing["tracks"] = final_tracks
                    new_pl = existing
                else:
                    new_pl = {
                        "id": secrets.token_urlsafe(6),
                        "name": auto_pl_name,
                        "tracks": final_tracks,
                    }
                    playlists.append(new_pl)
                user_manager.users[username]["music_playlists"] = playlists
                user_manager.save_users()
                _dl_status[did]["playlist"] = new_pl["name"]

            count = len(final_tracks)
            _dl_status[did].update({
                "status": "done", "total": count,
                "filename": f"{count} track(s) downloaded",
                "playlist": _dl_status[did].get("playlist"),
            })
            logger.info(f"Channel download done: {url} ({count} tracks), playlist: {auto_pl_name}")
        except Exception as e:
            _dl_status[did].update({"status": "error", "error": str(e)})
            logger.error(f"Channel download error: {e}")

    _threading.Thread(target=run, daemon=True).start()
    return jsonify({"success": True, "download_id": did}), 202


@app.route("/api/music/youtube/status/<did>", methods=["GET"])
@jwt_required()
def music_youtube_status(did):
    s = _dl_status.get(did)
    if not s:
        return jsonify({"msg": "Unknown ID"}), 404
    return jsonify(s), 200


@app.route("/api/music/channel/status/<did>", methods=["GET"])
@jwt_required()
def music_channel_status(did):
    s = _dl_status.get(did)
    if not s:
        return jsonify({"msg": "Unknown ID"}), 404
    return jsonify(s), 200


@app.route("/api/music/delete", methods=["DELETE"])
@jwt_required()
def music_delete():
    data = request.get_json() or {}
    fname = data.get("filename", "").strip()
    path  = _safe_music_path(fname) if fname else None
    if not path or not path.exists():
        return jsonify({"msg": "Not found"}), 404
    path.unlink()
    return jsonify({"msg": "Deleted"}), 200


@app.route("/api/music/rename", methods=["POST"])
@jwt_required()
def music_rename():
    data = request.get_json() or {}
    old  = data.get("old_name", "").strip()
    new  = data.get("new_name", "").strip()
    if not old or not new:
        return jsonify({"msg": "Both names required"}), 400
    old_path = _safe_music_path(old)
    if not old_path or not old_path.exists():
        return jsonify({"msg": "Not found"}), 404
    _, ext = os.path.splitext(old)
    if not os.path.splitext(new)[1]:
        new += ext
    safe_new = re.sub(r'[/\\:*?"<>|]', '_', new).strip()
    new_path = _safe_music_path(safe_new)
    if not new_path:
        return jsonify({"msg": "Invalid name"}), 400
    if new_path.exists():
        return jsonify({"msg": "Name already exists"}), 409
    old_path.rename(new_path)
    return jsonify({"msg": "Renamed", "new_name": new_path.name}), 200


@app.route("/api/music/prefs", methods=["GET"])
@jwt_required()
def music_get_prefs():
    username = get_jwt_identity()
    user = user_manager.get_user(username)
    return jsonify({
        "playlists":  user.get("music_playlists", []),
        "favorites":  user.get("music_favorites", []),
    }), 200


@app.route("/api/music/prefs", methods=["POST"])
@jwt_required()
def music_save_prefs():
    username = get_jwt_identity()
    data = request.get_json() or {}
    if "playlists" in data:
        user_manager.users[username]["music_playlists"] = data["playlists"]
    if "favorites" in data:
        user_manager.users[username]["music_favorites"] = data["favorites"]
    user_manager.save_users()
    return jsonify({"msg": "Saved"}), 200


# ──────────────────────────────────────────────────────────────────
#  ENTRY POINT  (must be last — after all routes are defined)
# ──────────────────────────────────────────────────────────────────


# ──────────────────────────────────────────────────────────────────
#  MUSIC SCAN — refresh library (picks up files dropped in /tracks/)
# ──────────────────────────────────────────────────────────────────

@app.route("/api/music/scan", methods=["POST"])
@jwt_required()
def music_scan():
    try:
        files = []
        for f in MUSIC_DIR.iterdir():
            if f.is_file() and f.suffix.lstrip('.').lower() in MUSIC_EXTENSIONS:
                st = f.stat()
                files.append({"name": f.name, "size": st.st_size,
                               "size_fmt": _music_size_fmt(st.st_size),
                               "modified": st.st_mtime})
        files.sort(key=lambda x: x["name"].lower())
        return jsonify({"tracks": files, "count": len(files)}), 200
    except Exception as e:
        return jsonify({"msg": str(e)}), 500


# ──────────────────────────────────────────────────────────────────
#  ENRICH — bulk artist lookup for existing tracks via MusicBrainz
#  Renames "Song Title.mp3" → "Artist - Song Title.mp3"
# ──────────────────────────────────────────────────────────────────

_enrich_status = {}   # job_id -> {done, total, renamed, errors}

def _mb_lookup(title, duration_sec=None):
    """Query MusicBrainz for artist given a track title. Returns artist str or None."""
    try:
        q = _urllib_parse.urlencode({"query": f'recording:"{title}"', "fmt": "json", "limit": "5"})
        req = _urllib_req.Request(
            f"https://musicbrainz.org/ws/2/recording?{q}",
            headers={"User-Agent": "NAS-MusicPlayer/1.0 (nas@localhost)"}
        )
        with _urllib_req.urlopen(req, timeout=6) as r:
            mb = json.loads(r.read())
        recordings = mb.get("recordings", [])
        if not recordings:
            return None
        best = recordings[0]
        if duration_sec:
            for rec in recordings:
                rec_dur = (rec.get("length") or 0) / 1000
                if abs(rec_dur - duration_sec) < 8:
                    best = rec; break
        credits = best.get("artist-credit", [])
        if credits:
            return credits[0].get("artist", {}).get("name")
    except Exception:
        pass
    return None

@app.route("/api/music/enrich", methods=["POST"])
@jwt_required()
def music_enrich():
    """Background job: find artist names for tracks that lack them."""
    import time as _time
    jid = secrets.token_urlsafe(6)
    _enrich_status[jid] = {"done": 0, "total": 0, "renamed": [], "errors": [], "running": True}

    def run():
        # Find tracks without an artist (no " - " separator in basename)
        candidates = []
        for f in MUSIC_DIR.iterdir():
            if f.is_file() and f.suffix.lstrip('.').lower() in MUSIC_EXTENSIONS:
                stem = f.stem
                if ' - ' not in stem:
                    candidates.append(f)
        _enrich_status[jid]["total"] = len(candidates)

        for f in candidates:
            stem = f.stem
            # Strip common noise
            clean = re.sub(
                r'\s*[\(\[].*?(official|video|audio|lyrics|hd|4k|mv|music|feat\.?|ft\.?).*?[\)\]]',
                '', stem, flags=re.IGNORECASE
            ).strip()
            artist = _mb_lookup(clean)
            _enrich_status[jid]["done"] += 1
            if not artist:
                _enrich_status[jid]["errors"].append(stem)
                continue
            new_stem = re.sub(r'[/\\:*?"<>|]', '_', f"{artist} - {clean}").strip()
            new_path = MUSIC_DIR / (new_stem + f.suffix)
            # Don't overwrite existing file
            if new_path.exists():
                _enrich_status[jid]["errors"].append(stem)
                continue
            try:
                f.rename(new_path)
                _enrich_status[jid]["renamed"].append({"old": f.name, "new": new_path.name})
            except Exception as ex:
                _enrich_status[jid]["errors"].append(f"{stem}: {ex}")
            _time.sleep(0.3)   # be polite to MusicBrainz API

        _enrich_status[jid]["running"] = False

    _threading.Thread(target=run, daemon=True).start()
    return jsonify({"job_id": jid}), 202


@app.route("/api/music/enrich/status/<jid>", methods=["GET"])
@jwt_required()
def music_enrich_status(jid):
    s = _enrich_status.get(jid)
    if not s:
        return jsonify({"msg": "Unknown job"}), 404
    return jsonify(s), 200


# ──────────────────────────────────────────────────────────────────
#  LYRICS — proxy to lrclib.net (free, no API key needed)
# ──────────────────────────────────────────────────────────────────

import urllib.request as _urllib_req
import urllib.parse   as _urllib_parse

@app.route("/api/music/lyrics", methods=["GET"])
@jwt_required()
def music_lyrics():
    track_name  = request.args.get("track_name",  "").strip()
    artist_name = request.args.get("artist_name", "").strip()
    if not track_name:
        return jsonify({"found": False, "msg": "track_name required"}), 400
    try:
        params = {"track_name": track_name}
        if artist_name:
            params["artist_name"] = artist_name
        url = "https://lrclib.net/api/search?" + _urllib_parse.urlencode(params)
        req = _urllib_req.Request(url, headers={"User-Agent": "NAS-Music/1.0"})
        with _urllib_req.urlopen(req, timeout=8) as r:
            results = json.loads(r.read().decode())
        if not results:
            return jsonify({"found": False}), 200
        best = results[0]
        synced_raw = best.get("syncedLyrics", "") or ""
        plain_raw  = best.get("plainLyrics",  "") or ""
        if synced_raw:
            return jsonify({"found": True, "synced": True,
                            "lines": _parse_lrc(synced_raw), "lyrics": plain_raw,
                            "title": best.get("trackName",""), "artist": best.get("artistName","")}), 200
        elif plain_raw:
            return jsonify({"found": True, "synced": False, "lines": [],
                            "lyrics": plain_raw,
                            "title": best.get("trackName",""), "artist": best.get("artistName","")}), 200
        return jsonify({"found": False}), 200
    except Exception as e:
        logger.error(f"Lyrics error: {e}")
        return jsonify({"found": False, "msg": str(e)}), 200


def _parse_lrc(lrc_text):
    import re as _re
    lines = []
    pat = _re.compile(r'\[(\d+):(\d+)\.(\d+)\](.*)')
    for line in lrc_text.splitlines():
        m = pat.match(line.strip())
        if m:
            mins, secs, cs, text = m.groups()
            ms = (int(mins)*60 + int(secs))*1000 + int(cs.ljust(3,"0")[:3])
            text = text.strip()
            if text:
                lines.append({"time_ms": ms, "text": text})
    return lines


if __name__ == "__main__":
    try:
        import PIL
        print("✓ Pillow installed")
    except ImportError:
        print("⚠️  Pillow not installed — run: pip install Pillow")

    try:
        import yt_dlp
        print("✓ yt-dlp installed")
    except ImportError:
        print("⚠️  yt-dlp not installed — run: pip install yt-dlp")

    print("\n" + "=" * 60)
    print("📦  NAS + MUSIC SYSTEM STARTING")
    print("=" * 60)
    print(f"  Storage : {STORAGE_DIR}")
    print(f"  Music   : {MUSIC_DIR}")
    print(f"  Static  : {STATIC_DIR}")
    print(f"  Logs    : {LOG_DIR}")
    print("\n  Default admin credentials:")
    print("    Username : admin")
    print("    Password : admin123")
    print("\n  ⚠️  CHANGE THE ADMIN PASSWORD AFTER FIRST LOGIN!")
    print("=" * 60 + "\n")

    debug_mode = os.getenv("FLASK_DEBUG", "false").lower() == "true"
    port       = int(os.getenv("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=debug_mode)
